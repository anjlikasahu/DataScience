
from openpyxl import load_workbook
import mysql.connector
import MySQLdb



file = r"C:\Users\Anjlika Sahu\OneDrive\Desktop\TECH\SARCI\Task1\sample.xlsx"
wb_obj = load_workbook(file)

sheet_obj = wb_obj.active
loc = r"C:\Users\Anjlika Sahu\OneDrive\Desktop\TECH\SARCI\Task1\abc.xlsx"
columns = list()  # for getting all the column names in one list
row_data = list()  # for getting all the row data onr by one
final = list()  # appending the row_data list content int the final matrix

# row 1 has all the column names so appending all the names one by one into the "column" list
max_column = sheet_obj.max_column
for i in range(1, max_column + 1):
    cell_obj = sheet_obj.cell(row=1, column=i)
    columns.append(cell_obj.value)

max_col = sheet_obj.max_column  # max_column = total number of columns
max_row = sheet_obj.max_row  # max_row = total number of rows
for i in range(2, max_row + 1):
    for j in range(1, max_col + 1):
        cell_obj2 = sheet_obj.cell(row=i, column=j)
        row_data.append(cell_obj2.value)
    final.append(row_data)
    row_data = list()


mydb = mysql.connector.connect(host = "localhost", user="Anj", passwd= "Anjii$2015", db= "pythondb1")
cursor = mydb.cursor()

query = """
           INSERT INTO orders_1 (`Level`, `Position`, `Customer Part Number`, `Description`, `Ware-house`, `Revision`,`Effect.Date`, 
           `Expiry Date`, `Locations`, `Location Qty`, `Scrap [%]`,`Scrap Quantity`, `Inv Unt`, `Net Quantity`) VALUES 
           (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
"""

for i in range(0, len(final)):
    Level = final[i][0]
    Position = final[i][1]
    Number = final[i][2]
    Desc = final[i][3]
    Ware_house = final[i][4]
    Revision = final[i][5]
    Effect_Date = final[i][6]
    Expiry_Date = final[i][7]
    Locations = final[i][8]
    Location_Qty = final[i][9]
    Scrap = final[i][10]
    Scrap_Quantity = final[i][11]
    Inv_Unt = final[i][12]
    Net_Quantity = final[i][13]

    values = (
    Level, Position, Number, Desc, Ware_house, Revision, Effect_Date, Expiry_Date, Locations, Location_Qty, Scrap,
    Scrap_Quantity, Inv_Unt, Net_Quantity)

    cursor.execute(query, values)

cursor.close()

mydb.commit()

mydb.close()

print(" ")
print("All done")
print(" ")

columns = str(len(final[0]))
rows = str(len(final))
print("Imported " + columns + " columns and " + rows + " rows to MySQL")



