from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

file = r"C:\Users\Anjlika Sahu\OneDrive\Desktop\TECH\SARCI\Task1\sample.xlsx"
wb_obj = load_workbook(file)
sheet_obj = wb_obj.active
loc = r"C:\Users\Anjlika Sahu\OneDrive\Desktop\TECH\SARCI\Task1\pypy4.xlsx"
columns = list()  # for getting all the column names in one list

row_data = list()  # for getting all the row data onr by one

final = list()  # appending the row_data list content int the final matrix

# row 1 has all the column names so appending all the names one by one into the "column" list
max_column = sheet_obj.max_column
for i in range(1, max_column + 1):
    cell_obj = sheet_obj.cell(row=1, column=i)
    columns.append(cell_obj.value)

max_col = sheet_obj.max_column  # max_column = total number of columns
max_row = sheet_obj.max_row  # max_row = total number of rows
for i in range(2, max_row + 1):
    for j in range(1, max_col + 1):
        cell_obj2 = sheet_obj.cell(row = i, column = j)
        row_data.append(cell_obj2.value)
    final.append(row_data)
    row_data = list()

print(max_row)

cell_first = 'A1'  # coordinates of the first cell
cell_last = sheet_obj.cell(row = i, column = j)  # cell_last points to the last cell
REF = cell_first + ':' + cell_last.coordinate  # cell_last.coordinate gives the name of the last cell

wb = Workbook()
ws = wb.active
ws.append(columns)

for row in final:
    ws.append(row)  # appending the data extracted int the "final" list to workbook.

tab = Table(displayName = 'Table1', ref = REF)
style = TableStyleInfo(name = "TableStyleMedium9", showRowStripes = True, showColumnStripes = True)
tab.tableStyleInfo = style
ws.add_table(tab)
wb.save(loc)
